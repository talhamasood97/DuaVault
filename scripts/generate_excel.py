#!/usr/bin/env python3
"""
generate_excel.py
Generates a scholar review Excel file for Dua Companion duas.
Output: /Users/mohdtalhamasood/Downloads/Duas_Scholar_Review.xlsx
"""

import subprocess
import sys

# ── Ensure openpyxl is available ─────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("openpyxl not found – installing…")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Dua data (extracted from data/duas.ts) ───────────────────────────────────
DUAS = [
    # id, slug, title, arabic_text, transliteration, translation,
    # source_book, hadith_number, authenticity_grade, category,
    # emotion_tags, situation_tags, scholar_verified, featured, daily_dua_eligible

    (1, "dua-before-sleeping", "Dua Before Sleeping",
     "بِاسْمِكَ اللَّهُمَّ أَمُوتُ وَأَحْيَا",
     "Bismika Allahumma amutu wa ahya",
     "In Your name, O Allah, I die and I live.",
     "Sahih al-Bukhari", "6324", "SAHIH", "daily-life",
     ["grateful", "hopeful", "lonely"],
     ["before-sleep", "night", "bedtime", "sleep", "sleeping"],
     True, True, True),

    (2, "dua-upon-waking", "Dua Upon Waking Up",
     "الْحَمْدُ لِلَّهِ الَّذِي أَحْيَانَا بَعْدَ مَا أَمَاتَنَا وَإِلَيْهِ النُّشُورُ",
     "Alhamdu lillahil-ladhi ahyana ba'da ma amatana wa ilayhin-nushur",
     "All praise is for Allah who gave us life after having taken it from us, and unto Him is the resurrection.",
     "Sahih al-Bukhari", "6325", "SAHIH", "daily-life",
     ["grateful", "lonely"],
     ["morning", "waking-up"],
     True, True, True),

    (3, "dua-before-eating", "Dua Before Eating",
     "بِسْمِ اللَّهِ",
     "Bismillah",
     "In the name of Allah.",
     "Sahih al-Bukhari", "5376", "SAHIH", "daily-life",
     ["grateful"],
     ["eating", "food", "meal", "water", "drink"],
     True, False, True),

    (4, "dua-after-eating", "Dua After Eating",
     "الْحَمْدُ لِلَّهِ الَّذِي أَطْعَمَنَا وَسَقَانَا وَجَعَلَنَا مُسْلِمِينَ",
     "Alhamdu lillahil-ladhi at'amana wa saqana wa ja'alana muslimin",
     "All praise is for Allah who fed us and gave us drink, and made us Muslims.",
     "Sunan Abu Dawud", "3850", "SAHIH", "daily-life",
     ["grateful"],
     ["after-eating", "food", "gratitude", "water", "drink", "drinking"],
     True, False, True),

    (5, "dua-entering-home", "Dua When Entering Home",
     "بِسْمِ اللَّهِ وَلَجْنَا وَبِسْمِ اللَّهِ خَرَجْنَا وَعَلَى اللَّهِ رَبِّنَا تَوَكَّلْنَا",
     "Bismillahi walajna, wa bismillahi kharajna, wa 'ala Allahi rabbina tawakkalna",
     "In the name of Allah we enter, in the name of Allah we leave, and upon our Lord we place our trust.",
     "Sunan Abu Dawud", "5096", "SAHIH", "daily-life",
     ["grateful", "hopeful"],
     ["home", "entering", "door"],
     True, False, True),

    (6, "dua-leaving-home", "Dua When Leaving Home",
     "بِسْمِ اللَّهِ تَوَكَّلْتُ عَلَى اللَّهِ وَلَا حَوْلَ وَلَا قُوَّةَ إِلَّا بِاللَّهِ",
     "Bismillahi, tawakkaltu 'ala Allah, wa la hawla wa la quwwata illa billah",
     "In the name of Allah, I place my trust in Allah, and there is no might nor power except with Allah.",
     "Sunan Abu Dawud", "5095", "SAHIH", "daily-life",
     ["hopeful"],
     ["leaving-home", "door", "protection"],
     True, False, True),

    (7, "dua-entering-bathroom", "Dua Entering the Bathroom",
     "اللَّهُمَّ إِنِّي أَعُوذُ بِكَ مِنَ الْخُبُثِ وَالْخَبَائِثِ",
     "Allahumma inni a'udhu bika minal-khubuthi wal-khaba'ith",
     "O Allah, I seek refuge in You from the male and female evil ones (devils).",
     "Sahih al-Bukhari", "142", "SAHIH", "daily-life",
     ["afraid"],
     ["bathroom", "toilet", "entering"],
     True, False, False),

    (8, "dua-after-wudu", "Dua After Wudu (Ablution)",
     "أَشْهَدُ أَنْ لَا إِلَهَ إِلَّا اللَّهُ وَحْدَهُ لَا شَرِيكَ لَهُ وَأَشْهَدُ أَنَّ مُحَمَّدًا عَبْدُهُ وَرَسُولُهُ",
     "Ashhadu an la ilaha illallahu wahdahu la sharika lah, wa ashhadu anna Muhammadan 'abduhu wa rasuluh",
     "I bear witness that there is no deity worthy of worship except Allah, alone, without partner, and I bear witness that Muhammad is His slave and messenger.",
     "Sahih Muslim", "234", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["wudu", "ablution", "purification", "wudhu", "water"],
     True, False, True),

    (9, "dua-entering-mosque", "Dua When Entering the Mosque",
     "اللَّهُمَّ افْتَحْ لِي أَبْوَابَ رَحْمَتِكَ",
     "Allahumma iftah li abwaba rahmatik",
     "O Allah, open the gates of Your mercy for me.",
     "Sahih Muslim", "713", "SAHIH", "worship",
     ["hopeful", "grateful"],
     ["mosque", "masjid", "entering", "prayer"],
     True, True, True),

    (10, "dua-leaving-mosque", "Dua When Leaving the Mosque",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ مِنْ فَضْلِكَ",
     "Allahumma inni as'aluka min fadlik",
     "O Allah, I ask You of Your bounty.",
     "Sahih Muslim", "713", "SAHIH", "worship",
     ["hopeful", "grateful"],
     ["mosque", "masjid", "leaving", "prayer"],
     True, False, True),

    (11, "dua-morning-protection", "Morning Dua for Protection",
     "اللَّهُمَّ بِكَ أَصْبَحْنَا وَبِكَ أَمْسَيْنَا وَبِكَ نَحْيَا وَبِكَ نَمُوتُ وَإِلَيْكَ النُّشُورُ",
     "Allahumma bika asbahna, wa bika amsayna, wa bika nahya, wa bika namutu, wa ilayka n-nushur",
     "O Allah, by You we enter the morning and by You we enter the evening, by You we live and by You we die, and to You is the resurrection.",
     "Sunan al-Tirmidhi", "3391", "SAHIH", "morning-evening",
     ["grateful", "hopeful", "lonely"],
     ["morning", "adhkar", "protection"],
     True, True, True),

    (12, "dua-evening-protection", "Evening Dua for Protection",
     "اللَّهُمَّ بِكَ أَمْسَيْنَا وَبِكَ أَصْبَحْنَا وَبِكَ نَحْيَا وَبِكَ نَمُوتُ وَإِلَيْكَ الْمَصِيرُ",
     "Allahumma bika amsayna, wa bika asbahna, wa bika nahya, wa bika namutu, wa ilaykal-masir",
     "O Allah, by You we enter the evening and by You we enter the morning, by You we live and by You we die, and to You is the final return.",
     "Sunan al-Tirmidhi", "3391", "SAHIH", "morning-evening",
     ["grateful", "lonely"],
     ["evening", "adhkar"],
     True, False, True),

    (13, "sayyidul-istighfar", "Sayyidul Istighfar – Master Dua of Forgiveness",
     "اللَّهُمَّ أَنْتَ رَبِّي لَا إِلَهَ إِلَّا أَنْتَ خَلَقْتَنِي وَأَنَا عَبْدُكَ وَأَنَا عَلَى عَهْدِكَ وَوَعْدِكَ مَا اسْتَطَعْتُ أَعُوذُ بِكَ مِنْ شَرِّ مَا صَنَعْتُ أَبُوءُ لَكَ بِنِعْمَتِكَ عَلَيَّ وَأَبُوءُ لَكَ بِذَنْبِي فَاغْفِرْ لِي فَإِنَّهُ لَا يَغْفِرُ الذُّنُوبَ إِلَّا أَنْتَ",
     "Allahumma anta Rabbi la ilaha illa anta, khalaqtani wa ana abduka, wa ana 'ala 'ahdika wa wa'dika mastata'tu, a'udhu bika min sharri ma sana'tu, abu'u laka bini'matika 'alayya, wa abu'u laka bidhanbi faghfir li, fa innahu la yaghfirudh-dhunuba illa anta",
     "O Allah, You are my Lord. There is no god but You. You created me and I am Your slave. I am keeping my covenant and my promise to You as much as I am able. I seek refuge in You from the evil that I have done. I acknowledge Your favor upon me and I acknowledge my sin. Forgive me, for there is none who forgives sins except You.",
     "Sahih al-Bukhari", "6306", "SAHIH", "forgiveness",
     ["seeking-forgiveness", "sad", "hopeful"],
     ["forgiveness", "morning", "evening", "sins", "repentance"],
     True, True, True),

    (14, "dua-for-anxiety-and-grief", "Dua for Anxiety and Grief",
     "اللَّهُمَّ إِنِّي أَعُوذُ بِكَ مِنَ الْهَمِّ وَالْحَزَنِ وَالْعَجْزِ وَالْكَسَلِ وَالْبُخْلِ وَالْجُبْنِ وَضَلَعِ الدَّيْنِ وَغَلَبَةِ الرِّجَالِ",
     "Allahumma inni a'udhu bika minal-hammi wal-hazani, wal-'ajzi wal-kasali, wal-bukhli wal-jubni, wa dhala'id-dayni wa ghalabatir-rijal",
     "O Allah, I seek refuge in You from anxiety and sorrow, weakness and laziness, miserliness and cowardice, the burden of debts and from being overpowered by men.",
     "Sahih al-Bukhari", "6369", "SAHIH", "protection",
     ["stressed", "sad", "afraid", "lonely"],
     ["anxiety", "stress", "grief", "worry", "depression"],
     True, True, True),

    (15, "dua-for-protection-morning", "Dua for Complete Protection (3x Morning)",
     "بِسْمِ اللَّهِ الَّذِي لَا يَضُرُّ مَعَ اسْمِهِ شَيْءٌ فِي الْأَرْضِ وَلَا فِي السَّمَاءِ وَهُوَ السَّمِيعُ الْعَلِيمُ",
     "Bismillahil-ladhi la yadurru ma'asmihi shay'un fil-ardi wa la fis-sama'i wa huwas-sami'ul-'alim",
     "In the name of Allah, with whose name nothing on earth or in heaven can cause harm, and He is the All-Hearing, the All-Knowing.",
     "Sunan Abu Dawud", "5088", "SAHIH", "protection",
     ["afraid", "hopeful"],
     ["protection", "morning", "evening", "ruqyah"],
     True, False, True),

    (16, "dua-against-evil-eye", "Dua Against the Evil Eye (Al-Ayn)",
     "أَعُوذُ بِكَلِمَاتِ اللَّهِ التَّامَّاتِ مِنْ شَرِّ مَا خَلَقَ",
     "A'udhu bi kalimatillahit-tammati min sharri ma khalaq",
     "I seek refuge in the perfect words of Allah from the evil of what He has created.",
     "Sahih Muslim", "2708", "SAHIH", "protection",
     ["afraid"],
     ["evil-eye", "protection", "ruqyah", "harm"],
     True, False, True),

    (17, "dua-when-afraid", "Dua When Afraid or in Fear",
     "حَسْبُنَا اللَّهُ وَنِعْمَ الْوَكِيلُ",
     "Hasbunallahu wa ni'mal-wakil",
     "Allah is sufficient for us, and He is the best disposer of affairs.",
     "Sahih al-Bukhari", "4563", "SAHIH", "protection",
     ["afraid", "stressed", "hopeful", "angry"],
     ["fear", "worry", "trust-in-allah"],
     True, False, True),

    (18, "dua-for-forgiveness-simple", "Dua for Forgiveness",
     "رَبِّ اغْفِرْ لِي وَتُبْ عَلَيَّ إِنَّكَ أَنْتَ التَّوَّابُ الرَّحِيمُ",
     "Rabbighfir li wa tub 'alayya innaka antat-tawwabur-rahim",
     "My Lord, forgive me and accept my repentance. Indeed You are the Oft-Returning, the Most Merciful.",
     "Sunan al-Tirmidhi", "3434", "SAHIH", "forgiveness",
     ["seeking-forgiveness", "sad", "hopeful"],
     ["forgiveness", "repentance", "sins"],
     True, False, True),

    (19, "dua-laylatul-qadr", "Dua for Laylatul Qadr (Night of Power)",
     "اللَّهُمَّ إِنَّكَ عَفُوٌّ تُحِبُّ الْعَفْوَ فَاعْفُ عَنِّي",
     "Allahumma innaka 'afuwwun tuhibbul-'afwa fa'fu 'anni",
     "O Allah, You are Pardoning and love pardon, so pardon me.",
     "Sunan al-Tirmidhi", "3513", "SAHIH", "forgiveness",
     ["seeking-forgiveness", "hopeful", "lonely"],
     ["laylatul-qadr", "ramadan", "forgiveness", "night-of-power"],
     True, True, True),

    (20, "dua-for-parents", "Dua for Parents",
     "رَّبِّ ارْحَمْهُمَا كَمَا رَبَّيَانِي صَغِيرًا",
     "Rabbir hamhuma kama rabbayani saghira",
     "My Lord, have mercy upon them as they raised me when I was young.",
     "Quran", "17:24", "QURAN", "family",
     ["grateful", "hopeful", "sad"],
     ["parents", "family", "mercy", "quranic", "mother", "father"],
     True, True, True),

    (21, "dua-for-travel", "Dua for Travelling",
     "سُبْحَانَ الَّذِي سَخَّرَ لَنَا هَذَا وَمَا كُنَّا لَهُ مُقْرِنِينَ وَإِنَّا إِلَى رَبِّنَا لَمُنقَلِبُونَ",
     "Subhanal-ladhi sakhkhara lana hadha wa ma kunna lahu muqrinin, wa inna ila rabbina lamunqalibun",
     "Glory be to the One who has subjected this to us, and we were not capable of doing so ourselves. And indeed, to our Lord we will surely return.",
     "Sunan Abu Dawud", "2602", "SAHIH", "travel",
     ["grateful", "hopeful"],
     ["travel", "vehicle", "journey", "car", "plane", "umrah", "hajj", "pilgrimage"],
     True, False, False),

    (22, "dua-returning-from-travel", "Dua When Returning From Travel",
     "آيِبُونَ تَائِبُونَ عَابِدُونَ لِرَبِّنَا حَامِدُونَ",
     "Ayibuna, ta'ibuna, 'abiduna, lirabbina hamidun",
     "We are returning, repenting, worshipping, and praising our Lord.",
     "Sahih al-Bukhari", "3085", "SAHIH", "travel",
     ["grateful"],
     ["travel", "returning", "journey", "umrah", "hajj", "pilgrimage"],
     True, False, False),

    (23, "dua-rabbana-atina", "Rabbana Atina – The Complete Dua",
     "رَبَّنَا آتِنَا فِي الدُّنْيَا حَسَنَةً وَفِي الْآخِرَةِ حَسَنَةً وَقِنَا عَذَابَ النَّارِ",
     "Rabbana atina fid-dunya hasanatan wa fil-akhirati hasanatan wa qina 'adhaban-nar",
     "Our Lord, give us good in this world and good in the Hereafter, and protect us from the punishment of the Fire.",
     "Quran", "2:201", "QURAN", "quranic",
     ["hopeful", "grateful"],
     ["general", "dunya", "akhirah", "protection", "quranic"],
     True, True, True),

    (24, "dua-of-prophet-yunus", "Dua of Prophet Yunus (Jonah)",
     "لَّا إِلَهَ إِلَّا أَنتَ سُبْحَانَكَ إِنِّي كُنتُ مِنَ الظَّالِمِينَ",
     "La ilaha illa anta subhanaka inni kuntu minaz-zalimin",
     "There is no deity except You; exalted are You. Indeed, I have been of the wrongdoers.",
     "Quran", "21:87", "QURAN", "quranic",
     ["seeking-forgiveness", "sad", "hopeful", "stressed", "angry", "lonely"],
     ["distress", "hardship", "forgiveness", "prophets", "quranic"],
     True, True, True),

    (25, "dua-of-prophet-ibrahim", "Dua of Prophet Ibrahim (Abraham)",
     "رَبِّ اجْعَلْنِي مُقِيمَ الصَّلَاةِ وَمِن ذُرِّيَّتِي رَبَّنَا وَتَقَبَّلْ دُعَاءِ",
     "Rabbi j'alni muqimas-salati wa min dhurriyyati rabbana wa taqabbal du'a",
     "My Lord, make me an establisher of prayer, and from my descendants. Our Lord, and accept my supplication.",
     "Quran", "14:40", "QURAN", "quranic",
     ["hopeful", "grateful"],
     ["prayer", "family", "children", "quranic"],
     True, False, True),

    (26, "dua-of-prophet-ayyub", "Dua of Prophet Ayyub (Job)",
     "أَنِّي مَسَّنِيَ الضُّرُّ وَأَنتَ أَرْحَمُ الرَّاحِمِينَ",
     "Anni massaniyad-durru wa anta arhamur-rahimin",
     "Indeed, adversity has touched me, and you are the Most Merciful of the merciful.",
     "Quran", "21:83", "QURAN", "quranic",
     ["sad", "stressed", "hopeful", "afraid", "angry", "lonely"],
     ["illness", "hardship", "suffering", "patience", "quranic"],
     True, False, True),

    (27, "dua-hasbi-allah", "Dua of Reliance – Hasbiyallah",
     "حَسْبِيَ اللَّهُ لَا إِلَهَ إِلَّا هُوَ عَلَيْهِ تَوَكَّلْتُ وَهُوَ رَبُّ الْعَرْشِ الْعَظِيمِ",
     "Hasbiyallahu la ilaha illa huwa 'alayhi tawakkaltu wa huwa rabbul-'arshil-'azim",
     "Allah is sufficient for me. There is no deity except Him. Upon Him I rely and He is the Lord of the Great Throne.",
     "Quran", "9:129", "QURAN", "quranic",
     ["afraid", "stressed", "hopeful", "angry", "lonely"],
     ["reliance", "trust", "difficulty", "anxiety", "quranic"],
     True, False, True),

    (28, "dua-when-ill", "Dua When Sick or Ill",
     "اللَّهُمَّ رَبَّ النَّاسِ أَذْهِبِ الْبَأْسَ اشْفِ أَنْتَ الشَّافِي لَا شِفَاءَ إِلَّا شِفَاؤُكَ شِفَاءً لَا يُغَادِرُ سَقَمًا",
     "Allahumma Rabban-nasi, adhhibil-ba'sa, ishfi antash-shafi, la shifa'a illa shifa'uk, shifa'an la yughadiru saqama",
     "O Allah, Lord of mankind, remove the harm. Heal, for You are the Healer. There is no healing except Your healing, a healing that leaves no illness.",
     "Sahih al-Bukhari", "5743", "SAHIH", "health",
     ["sad", "afraid", "hopeful"],
     ["illness", "sickness", "healing", "ruqyah"],
     True, False, False),

    (29, "dua-visiting-sick", "Dua When Visiting a Sick Person",
     "لَا بَأْسَ طَهُورٌ إِنْ شَاءَ اللَّهُ",
     "La ba'sa tahurun in sha Allah",
     "No harm will come to you. It will be a purification, if Allah wills.",
     "Sahih al-Bukhari", "3616", "SAHIH", "health",
     ["sad", "hopeful"],
     ["visiting-sick", "hospital", "illness"],
     True, False, False),

    (30, "dua-ruqyah-self", "Dua for Self-Ruqyah (Healing Recitation)",
     "بِسْمِ اللَّهِ أَرْقِيكَ مِنْ كُلِّ شَيْءٍ يُؤْذِيكَ مِنْ شَرِّ كُلِّ نَفْسٍ أَوْ عَيْنٍ حَاسِدٍ اللَّهُ يَشْفِيكَ بِسْمِ اللَّهِ أَرْقِيكَ",
     "Bismillahi arqika min kulli shay'in yu'dhika, min sharri kulli nafsin aw 'aynin hasidin, Allahu yashfika, bismillahi arqika",
     "In the name of Allah I perform ruqyah for you, from everything that harms you, from the evil of every soul or envious eye. Allah heals you; in the name of Allah I perform ruqyah for you.",
     "Sahih Muslim", "2186", "SAHIH", "health",
     ["afraid", "hopeful"],
     ["ruqyah", "healing", "evil-eye", "protection"],
     True, False, False),

    (31, "dua-after-adhan", "Dua After the Adhan (Call to Prayer)",
     "اللَّهُمَّ رَبَّ هَذِهِ الدَّعْوَةِ التَّامَّةِ وَالصَّلَاةِ الْقَائِمَةِ آتِ مُحَمَّدًا الْوَسِيلَةَ وَالْفَضِيلَةَ وَابْعَثْهُ مَقَامًا مَحْمُودًا الَّذِي وَعَدْتَهُ",
     "Allahumma rabba hadhihid-da'watit-tammati was-salatil-qa'imati, ati Muhammadanil-wasilata wal-fadilata, wab'athhu maqaman mahmudanil-ladhi wa'adtah",
     "O Allah, Lord of this perfect call and established prayer, grant Muhammad the intercession and the favor, and raise him to the praiseworthy position You have promised him.",
     "Sahih al-Bukhari", "614", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["adhan", "prayer", "salah"],
     True, False, True),

    (32, "dua-istikhara", "Dua al-Istikhara (Seeking Guidance)",
     "اللَّهُمَّ إِنِّي أَسْتَخِيرُكَ بِعِلْمِكَ وَأَسْتَقْدِرُكَ بِقُدْرَتِكَ وَأَسْأَلُكَ مِنْ فَضْلِكَ الْعَظِيمِ فَإِنَّكَ تَقْدِرُ وَلَا أَقْدِرُ وَتَعْلَمُ وَلَا أَعْلَمُ وَأَنْتَ عَلَّامُ الْغُيُوبِ",
     "Allahumma inni astakhiruka bi'ilmika wa astaqdiruka biqudratika wa as'aluka min fadlikal-'azim, fa innaka taqdiru wa la aqdiru, wa ta'lamu wa la a'lamu, wa anta 'allamul-ghuyub",
     "O Allah, I seek Your guidance through Your knowledge, and I seek Your capability through Your power, and I ask You from Your great bounty. For surely You have power and I do not, and You know and I do not, and You know the unseen.",
     "Sahih al-Bukhari", "1166", "SAHIH", "worship",
     ["hopeful", "stressed"],
     ["istikhara", "decision", "guidance", "marriage", "job"],
     True, True, False),

    (33, "dua-breaking-fast", "Dua for Breaking the Fast (Iftar)",
     "اللَّهُمَّ لَكَ صُمْتُ وَبِكَ آمَنْتُ وَعَلَيْكَ تَوَكَّلْتُ وَعَلَى رِزْقِكَ أَفْطَرْتُ",
     "Allahumma laka sumtu wa bika amantu wa 'alayka tawakkaltu wa 'ala rizqika aftartu",
     "O Allah, for You I fasted, and in You I believe, and upon You I rely, and with Your provision I break my fast.",
     "Sunan Abu Dawud", "2358", "HASAN", "worship",
     ["grateful"],
     ["iftar", "ramadan", "fasting", "breaking-fast"],
     True, False, False),

    (34, "dua-entering-ramadan", "Dua for Welcoming Ramadan",
     "اللَّهُمَّ بَارِكْ لَنَا فِي رَجَبٍ وَشَعْبَانَ وَبَلِّغْنَا رَمَضَانَ",
     "Allahumma barik lana fi Rajaba wa Sha'bana wa ballighna Ramadan",
     "O Allah, bless us in Rajab and Sha'ban, and allow us to reach Ramadan.",
     "Musnad Ahmad", "2346", "HASAN", "worship",
     ["hopeful", "grateful"],
     ["ramadan", "rajab", "shaban", "months"],
     True, False, False),

    (35, "dua-for-righteous-children", "Dua for Righteous Children",
     "رَبِّ هَبْ لِي مِنَ الصَّالِحِينَ",
     "Rabbi hab li minas-salihin",
     "My Lord, grant me from among the righteous.",
     "Quran", "37:100", "QURAN", "family",
     ["hopeful"],
     ["children", "family", "offspring", "quranic"],
     True, False, True),

    (36, "dua-for-righteous-spouse", "Dua for a Righteous Spouse",
     "رَبَّنَا هَبْ لَنَا مِنْ أَزْوَاجِنَا وَذُرِّيَّاتِنَا قُرَّةَ أَعْيُنٍ وَاجْعَلْنَا لِلْمُتَّقِينَ إِمَامًا",
     "Rabbana hab lana min azwajina wa dhurriyyatina qurrata a'yunin waj'alna lil-muttaqina imama",
     "Our Lord, grant us from among our wives and offspring comfort to our eyes, and make us a leader for the righteous.",
     "Quran", "25:74", "QURAN", "family",
     ["hopeful", "lonely"],
     ["marriage", "spouse", "family", "children", "quranic"],
     True, False, True),

    (37, "dua-for-deceased", "Dua for the Deceased",
     "اللَّهُمَّ اغْفِرْ لَهُ وَارْحَمْهُ وَعَافِهِ وَاعْفُ عَنْهُ",
     "Allahummaghfir lahu warhamhu wa 'afihi wa'fu 'anhu",
     "O Allah, forgive him, have mercy on him, pardon him, and excuse him.",
     "Sahih Muslim", "963", "SAHIH", "family",
     ["sad", "hopeful"],
     ["funeral", "deceased", "death", "janazah"],
     True, False, False),

    (38, "dua-for-depression-and-sadness", "Dua for Depression and Sadness",
     "اللَّهُمَّ إِنِّي عَبْدُكَ وَابْنُ عَبْدِكَ وَابْنُ أَمَتِكَ نَاصِيَتِي بِيَدِكَ مَاضٍ فِيَّ حُكْمُكَ عَدْلٌ فِيَّ قَضَاؤُكَ أَسْأَلُكَ بِكُلِّ اسْمٍ هُوَ لَكَ سَمَّيْتَ بِهِ نَفْسَكَ أَوْ أَنْزَلْتَهُ فِي كِتَابِكَ أَوْ عَلَّمْتَهُ أَحَدًا مِنْ خَلْقِكَ أَوِ اسْتَأْثَرْتَ بِهِ فِي عِلْمِ الْغَيْبِ عِنْدَكَ أَنْ تَجْعَلَ الْقُرْآنَ رَبِيعَ قَلْبِي وَنُورَ صَدْرِي وَجَلَاءَ حُزْنِي وَذَهَابَ هَمِّي",
     "Allahumma inni 'abduka, wabnu 'abdika, wabnu amatika, nasiyati biyadika, madin fiyya hukmuka, 'adlun fiyya qada'uka, as'aluka bi kulli ismin huwa laka, sammayta bihi nafsaka, aw anzaltahu fi kitabika, aw 'allamtahu ahadan min khalqika, awista'tharta bihi fi 'ilmil-ghaybi 'indaka, an taj'alal-Qur'ana rabi'a qalbi, wa nura sadri, wa jala'a huzni, wa dhahaba hammi",
     "O Allah, I am Your slave, son of Your slave and son of Your maidservant. My forelock is in Your hand, Your judgment is executed upon me, Your decree is just. I ask You by every name which belongs to You, by which You named Yourself, or revealed in Your Book, or taught to anyone of Your creation, or kept to Yourself in the knowledge of the unseen with You, to make the Quran the spring of my heart, the light of my chest, the remover of my sadness and the reliever of my distress.",
     "Musnad Ahmad", "3712", "SAHIH", "protection",
     ["sad", "stressed", "lonely", "hopeful", "angry"],
     ["depression", "sadness", "distress", "anxiety", "healing"],
     True, True, True),

    (39, "dua-for-gratitude", "Dua of Gratitude (Shukr)",
     "اللَّهُمَّ أَعِنِّي عَلَى ذِكْرِكَ وَشُكْرِكَ وَحُسْنِ عِبَادَتِكَ",
     "Allahumma a'inni 'ala dhikrika wa shukrika wa husni 'ibadatik",
     "O Allah, help me to remember You, to be grateful to You, and to worship You well.",
     "Sunan Abu Dawud", "1522", "SAHIH", "worship",
     ["grateful"],
     ["gratitude", "dhikr", "worship", "zikr"],
     True, False, True),

    (40, "dua-when-angry", "Dua When Feeling Angry",
     "أَعُوذُ بِاللَّهِ مِنَ الشَّيْطَانِ الرَّجِيمِ",
     "A'udhu billahi minash-shaytanir-rajim",
     "I seek refuge in Allah from the accursed Satan.",
     "Sahih al-Bukhari", "6115", "SAHIH", "protection",
     ["angry"],
     ["anger", "conflict", "shaytan"],
     True, False, True),

    (41, "dua-for-knowledge", "Dua for Knowledge and Wisdom",
     "رَّبِّ زِدْنِي عِلْمًا",
     "Rabbi zidni 'ilma",
     "My Lord, increase me in knowledge.",
     "Quran", "20:114", "QURAN", "quranic",
     ["hopeful"],
     ["knowledge", "studying", "wisdom", "exam", "quranic"],
     True, False, True),

    (42, "dua-for-ease-in-hardship", "Dua for Ease in Difficulty",
     "اللَّهُمَّ لَا سَهْلَ إِلَّا مَا جَعَلْتَهُ سَهْلًا وَأَنْتَ تَجْعَلُ الْحَزْنَ إِذَا شِئْتَ سَهْلًا",
     "Allahumma la sahla illa ma ja'altahu sahlan, wa anta taj'alul-hazna idha shi'ta sahla",
     "O Allah, there is no ease except that which You make easy, and You make the difficult easy if You wish.",
     "Ibn Hibban", "2427", "SAHIH", "protection",
     ["stressed", "sad", "afraid", "hopeful", "angry", "lonely"],
     ["hardship", "ease", "difficulty", "problem"],
     True, False, True),

    (43, "dua-for-the-night-of-jumu-ah", "Dua for Friday (Jumu'ah)",
     "اللَّهُمَّ صَلِّ عَلَى مُحَمَّدٍ وَعَلَى آلِ مُحَمَّدٍ",
     "Allahumma salli 'ala Muhammadin wa 'ala ali Muhammad",
     "O Allah, send blessings upon Muhammad and upon the family of Muhammad.",
     "Sahih al-Bukhari", "3370", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["friday", "jumu'ah", "salawat", "prophet", "jummah", "jumah"],
     True, False, True),

    (44, "dua-when-it-rains", "Dua When It Rains",
     "اللَّهُمَّ صَيِّبًا نَافِعًا",
     "Allahumma sayyiban nafi'a",
     "O Allah, let it be a beneficial rain.",
     "Sahih al-Bukhari", "1032", "SAHIH", "daily-life",
     ["grateful"],
     ["rain", "weather"],
     True, False, False),

    (45, "dua-for-guidance", "Dua for Guidance on the Straight Path",
     "اهْدِنَا الصِّرَاطَ الْمُسْتَقِيمَ صِرَاطَ الَّذِينَ أَنْعَمْتَ عَلَيْهِمْ غَيْرِ الْمَغْضُوبِ عَلَيْهِمْ وَلَا الضَّالِّينَ",
     "Ihdinas-siratal-mustaqim, siratal-ladhina an'amta 'alayhim, ghayril-maghdubi 'alayhim wa lad-dallin",
     "Guide us to the straight path – the path of those upon whom You have bestowed favor, not of those who have evoked anger or of those who are astray.",
     "Quran", "1:6-7", "QURAN", "quranic",
     ["hopeful", "seeking-forgiveness"],
     ["guidance", "straight-path", "daily", "quranic", "fatiha"],
     True, True, True),

    (46, "dua-for-steadfastness", "Dua for Steadfastness in Faith",
     "يَا مُقَلِّبَ الْقُلُوبِ ثَبِّتْ قَلْبِي عَلَى دِينِكَ",
     "Ya muqallibal-qulub, thabbit qalbi 'ala dinik",
     "O Turner of hearts, keep my heart firm on Your religion.",
     "Sunan al-Tirmidhi", "2140", "SAHIH", "protection",
     ["afraid", "hopeful", "seeking-forgiveness"],
     ["faith", "steadfastness", "heart", "iman"],
     True, False, True),

    (47, "dua-when-in-debt", "Dua When in Debt or Financial Hardship",
     "اللَّهُمَّ اكْفِنِي بِحَلَالِكَ عَنْ حَرَامِكَ وَأَغْنِنِي بِفَضْلِكَ عَمَّنْ سِوَاكَ",
     "Allahumma-kfini bihalilika 'an haramika wa aghnini bifadlika 'amman siwak",
     "O Allah, suffice me with Your lawful against Your prohibited, and make me independent of all others by Your grace.",
     "Sunan al-Tirmidhi", "3563", "HASAN", "daily-life",
     ["stressed", "sad", "hopeful"],
     ["debt", "financial", "poverty", "money", "rizq"],
     True, False, True),

    (48, "dua-for-patience", "Dua for Patience",
     "رَبَّنَا أَفْرِغْ عَلَيْنَا صَبْرًا وَثَبِّتْ أَقْدَامَنَا وَانصُرْنَا عَلَى الْقَوْمِ الْكَافِرِينَ",
     "Rabbana afrigh 'alayna sabran wa thabbit aqdamana wansurna 'alal-qawmil-kafirin",
     "Our Lord, pour upon us patience and plant firmly our feet and give us victory over the disbelieving people.",
     "Quran", "2:250", "QURAN", "quranic",
     ["stressed", "afraid", "hopeful", "angry"],
     ["patience", "sabr", "difficulty", "steadfastness", "quranic"],
     True, False, True),

    (49, "dua-for-good-health", "Dua for Good Health and Wellbeing",
     "اللَّهُمَّ عَافِنِي فِي بَدَنِي اللَّهُمَّ عَافِنِي فِي سَمْعِي اللَّهُمَّ عَافِنِي فِي بَصَرِي",
     "Allahumma 'afini fi badani, Allahumma 'afini fi sam'i, Allahumma 'afini fi basari",
     "O Allah, grant me health in my body. O Allah, grant me health in my hearing. O Allah, grant me health in my sight.",
     "Sunan Abu Dawud", "5090", "HASAN", "health",
     ["grateful", "hopeful"],
     ["health", "body", "wellbeing", "healing"],
     True, False, True),

    (50, "dua-tawakkul-reliance", "Dua of Complete Reliance on Allah",
     "تَوَكَّلْتُ عَلَى الْحَيِّ الَّذِي لَا يَمُوتُ",
     "Tawakkaltu 'alal-hayyil-ladhi la yamut",
     "I place my trust in the Ever-Living who does not die.",
     "Mustadrak al-Hakim", "1/689", "HASAN", "protection",
     ["afraid", "stressed", "hopeful"],
     ["reliance", "trust", "tawakkul", "protection"],
     True, False, True),

    (51, "dua-istiftah-opening-prayer", "Dua al-Istiftah – Opening Supplication of Prayer",
     "سُبْحَانَكَ اللَّهُمَّ وَبِحَمْدِكَ وَتَبَارَكَ اسْمُكَ وَتَعَالَى جَدُّكَ وَلَا إِلَهَ غَيْرُكَ",
     "Subhanakallahumma wa bihamdika wa tabarakasmuka wa ta'ala jadduka wa la ilaha ghairuk",
     "Glory and praise be to You, O Allah. Blessed is Your name and exalted is Your majesty. There is no deity worthy of worship except You.",
     "Sunan Abu Dawud", "775", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["prayer", "salah", "opening", "istiftah"],
     True, False, True),

    (52, "dua-in-ruku", "Dua in Ruku (Bowing in Prayer)",
     "سُبْحَانَ رَبِّيَ الْعَظِيمِ",
     "Subhana Rabbiyal-Azim",
     "How perfect is my Lord, the Most Great.",
     "Sahih Muslim", "772", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["prayer", "salah", "ruku", "bowing"],
     True, False, False),

    (53, "dua-in-sujood", "Dua in Sujood (Prostration)",
     "سُبْحَانَ رَبِّيَ الْأَعْلَى",
     "Subhana Rabbiyal-A'la",
     "How perfect is my Lord, the Most High.",
     "Sahih Muslim", "772", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["prayer", "salah", "sujood", "prostration"],
     True, False, False),

    (54, "dua-qunoot-witr", "Dua Qunoot for Witr Prayer",
     "اللَّهُمَّ اهْدِنِي فِيمَنْ هَدَيْتَ وَعَافِنِي فِيمَنْ عَافَيْتَ وَتَوَلَّنِي فِيمَنْ تَوَلَّيْتَ وَبَارِكْ لِي فِيمَا أَعْطَيْتَ وَقِنِي شَرَّ مَا قَضَيْتَ فَإِنَّكَ تَقْضِي وَلَا يُقْضَى عَلَيْكَ وَإِنَّهُ لَا يَذِلُّ مَنْ وَالَيْتَ تَبَارَكْتَ رَبَّنَا وَتَعَالَيْتَ",
     "Allahumma-hdini fiman hadayt, wa 'afini fiman 'afayt, wa tawallani fiman tawallayt, wa barik li fima a'tayt, wa qini sharra ma qadayt, fa innaka taqdi wa la yuqda 'alayk, wa innahu la yadhillu man walayt, tabarakta Rabbana wa ta'alayt",
     "O Allah, guide me among those You have guided, pardon me among those You have pardoned, befriend me among those You have befriended, bless me in what You have given, and protect me from the evil of what You have decreed. Surely You decree and none can decree over You. Truly the one whom You support is not humbled. Blessed and exalted are You, our Lord.",
     "Sunan al-Tirmidhi", "464", "SAHIH", "worship",
     ["hopeful", "grateful", "lonely"],
     ["witr", "qunoot", "night-prayer", "prayer"],
     True, False, True),

    (55, "dua-on-day-of-arafah", "Best Dua on the Day of Arafah",
     "لَا إِلَهَ إِلَّا اللَّهُ وَحْدَهُ لَا شَرِيكَ لَهُ لَهُ الْمُلْكُ وَلَهُ الْحَمْدُ يُحْيِي وَيُمِيتُ وَهُوَ عَلَى كُلِّ شَيْءٍ قَدِيرٌ",
     "La ilaha illallahu wahdahu la sharika lahu, lahul-mulku wa lahul-hamdu yuhyi wa yumitu wa huwa 'ala kulli shay'in qadir",
     "There is no deity worthy of worship except Allah, alone, without any partner. To Him belongs all dominion and all praise. He gives life and causes death, and He has power over all things.",
     "Sunan al-Tirmidhi", "3585", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["arafah", "hajj", "dhul-hijjah", "dhikr", "umrah", "pilgrimage", "arafat"],
     True, False, True),

    (56, "dua-after-prayer-dhikr", "Dhikr After Every Prayer",
     "سُبْحَانَ اللَّهِ وَالْحَمْدُ لِلَّهِ وَاللَّهُ أَكْبَرُ",
     "Subhanallahi walhamdu lillahi wallahu akbar",
     "How perfect is Allah, all praise is for Allah, and Allah is the Greatest. (Recited 33 times each after every prayer.)",
     "Sahih Muslim", "597", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["after-prayer", "salah", "tasbih", "dhikr", "zikr"],
     True, False, True),

    (57, "dua-seeking-allahs-love", "Dua for Allah's Love",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ حُبَّكَ وَحُبَّ مَنْ يُحِبُّكَ وَحُبَّ عَمَلٍ يُقَرِّبُنِي إِلَى حُبِّكَ",
     "Allahumma inni as'aluka hubbak, wa hubba man yuhibbuk, wa hubba 'amalin yuqarribuni ila hubbik",
     "O Allah, I ask You for Your love, and the love of those who love You, and love of deeds that bring me closer to Your love.",
     "Sunan al-Tirmidhi", "3490", "HASAN", "worship",
     ["hopeful", "grateful"],
     ["love", "closeness-to-allah", "dhikr", "worship", "zikr"],
     True, False, True),

    (58, "dua-for-jannah", "Dua Asking for Jannah (Paradise)",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ الْجَنَّةَ وَأَعُوذُ بِكَ مِنَ النَّارِ",
     "Allahumma inni as'alukal-jannah wa a'udhu bika minan-nar",
     "O Allah, I ask You for Paradise and I seek refuge in You from the Fire.",
     "Sunan Abu Dawud", "792", "SAHIH", "worship",
     ["hopeful", "afraid"],
     ["jannah", "paradise", "hellfire", "akhirah", "prayer"],
     True, False, True),

    (59, "ayatul-kursi", "Ayatul Kursi – The Throne Verse",
     "اللَّهُ لَا إِلَهَ إِلَّا هُوَ الْحَيُّ الْقَيُّومُ لَا تَأْخُذُهُ سِنَةٌ وَلَا نَوْمٌ لَّهُ مَا فِي السَّمَاوَاتِ وَمَا فِي الْأَرْضِ مَن ذَا الَّذِي يَشْفَعُ عِندَهُ إِلَّا بِإِذْنِهِ يَعْلَمُ مَا بَيْنَ أَيْدِيهِمْ وَمَا خَلْفَهُمْ وَلَا يُحِيطُونَ بِشَيْءٍ مِّنْ عِلْمِهِ إِلَّا بِمَا شَاءَ وَسِعَ كُرْسِيُّهُ السَّمَاوَاتِ وَالْأَرْضَ وَلَا يَئُودُهُ حِفْظُهُمَا وَهُوَ الْعَلِيُّ الْعَظِيمُ",
     "Allahu la ilaha illa huwal-hayyul-qayyum, la ta'khudhuhu sinatun wa la nawm, lahu ma fis-samawati wa ma fil-ard, man dhal-ladhi yashfa'u 'indahu illa bi-idhnih, ya'lamu ma bayna aydihim wa ma khalfahum, wa la yuhituna bi shay'in min 'ilmihi illa bima sha', wasi'a kursiyyuhus-samawati wal-ard, wa la ya'uduhu hifzuhuma wa huwal-'aliyyul-'azim",
     "Allah – there is no deity except Him, the Ever-Living, the Sustainer of existence. Neither drowsiness nor sleep overtakes Him. To Him belongs whatever is in the heavens and whatever is on the earth. Who can intercede with Him except by His permission? He knows what is before them and what will be after them, and they encompass nothing of His knowledge except what He wills. His Kursi extends over the heavens and the earth, and their preservation tires Him not. And He is the Most High, the Most Great.",
     "Quran", "2:255", "QURAN", "quranic",
     ["afraid", "hopeful", "grateful"],
     ["protection", "morning", "evening", "sleep", "ruqyah", "quranic"],
     True, True, True),

    (60, "dua-musa-ease-and-speech", "Dua of Prophet Musa for Ease and Clarity",
     "رَبِّ اشْرَحْ لِي صَدْرِي وَيَسِّرْ لِي أَمْرِي وَاحْلُلْ عُقْدَةً مِّن لِّسَانِي يَفْقَهُوا قَوْلِي",
     "Rabbi ishrah li sadri wa yassir li amri wahlul 'uqdatan min lisani yafqahu qawli",
     "My Lord, expand for me my chest, ease for me my task, and untie the knot from my tongue, that they may understand my speech.",
     "Quran", "20:25-28", "QURAN", "quranic",
     ["stressed", "hopeful"],
     ["exam", "speech", "confidence", "ease", "presentation", "quranic"],
     True, False, True),

    (61, "dua-of-zakariyya-for-offspring", "Dua of Prophet Zakariyya for Offspring",
     "رَبِّ لَا تَذَرْنِي فَرْدًا وَأَنتَ خَيْرُ الْوَارِثِينَ",
     "Rabbi la tadharni fardan wa anta khayrul-waritheen",
     "My Lord, do not leave me alone, and You are the best of inheritors.",
     "Quran", "21:89", "QURAN", "quranic",
     ["lonely", "hopeful"],
     ["children", "family", "loneliness", "offspring", "quranic"],
     True, False, True),

    (62, "dua-rabbana-la-tuzigh", "Dua Against Deviation of the Heart",
     "رَبَّنَا لَا تُزِغْ قُلُوبَنَا بَعْدَ إِذْ هَدَيْتَنَا وَهَبْ لَنَا مِن لَّدُنكَ رَحْمَةً إِنَّكَ أَنتَ الْوَهَّابُ",
     "Rabbana la tuzigh qulubana ba'da idh hadaytana wa hab lana min ladunka rahmah, innaka antal-wahhab",
     "Our Lord, do not let our hearts deviate after You have guided us, and grant us from Yourself mercy. Indeed, You are the Bestower.",
     "Quran", "3:8", "QURAN", "quranic",
     ["afraid", "hopeful", "seeking-forgiveness"],
     ["guidance", "heart", "iman", "faith", "quranic"],
     True, False, True),

    (63, "dua-rabbana-la-tu-akhidhna", "Dua of Forgiveness – Last Verse of Al-Baqarah",
     "رَبَّنَا لَا تُؤَاخِذْنَا إِن نَّسِينَا أَوْ أَخْطَأْنَا رَبَّنَا وَلَا تَحْمِلْ عَلَيْنَا إِصْرًا كَمَا حَمَلْتَهُ عَلَى الَّذِينَ مِن قَبْلِنَا رَبَّنَا وَلَا تُحَمِّلْنَا مَا لَا طَاقَةَ لَنَا بِهِ وَاعْفُ عَنَّا وَاغْفِرْ لَنَا وَارْحَمْنَا أَنتَ مَوْلَانَا فَانصُرْنَا عَلَى الْقَوْمِ الْكَافِرِينَ",
     "Rabbana la tu'akhidhna in nasina aw akhta'na, rabbana wa la tahmil 'alayna isran kama hamaltahu 'alal-ladhina min qablina, rabbana wa la tuhammilna ma la taqata lana bih, wa'fu 'anna waghfir lana warhamna, anta mawlana fansurna 'alal-qawmil-kafirin",
     "Our Lord, do not impose blame upon us if we have forgotten or erred. Our Lord, lay not upon us a burden like that which You laid upon those before us. Our Lord, burden us not with what we have no ability to bear. And pardon us; and forgive us; and have mercy upon us. You are our Protector, so give us victory over the disbelieving people.",
     "Quran", "2:286", "QURAN", "quranic",
     ["seeking-forgiveness", "stressed", "hopeful", "lonely"],
     ["forgiveness", "hardship", "burden", "daily", "quranic"],
     True, True, True),

    (64, "dua-rabbana-ghfir-li-walidayya", "Dua of Ibrahim for Parents and Believers",
     "رَبَّنَا اغْفِرْ لِي وَلِوَالِدَيَّ وَلِلْمُؤْمِنِينَ يَوْمَ يَقُومُ الْحِسَابُ",
     "Rabbana-ghfir li wa li walidayya wa lil-mu'minina yawma yaqumul-hisab",
     "Our Lord, forgive me and my parents and the believers on the Day when the reckoning will take place.",
     "Quran", "14:41", "QURAN", "family",
     ["hopeful", "seeking-forgiveness"],
     ["parents", "family", "forgiveness", "judgment-day", "quranic", "mother", "father"],
     True, False, True),

    (65, "dua-when-wearing-new-clothes", "Dua When Wearing New Clothes",
     "اللَّهُمَّ لَكَ الْحَمْدُ أَنْتَ كَسَوْتَنِيهِ أَسْأَلُكَ مِنْ خَيْرِهِ وَخَيْرِ مَا صُنِعَ لَهُ وَأَعُوذُ بِكَ مِنْ شَرِّهِ وَشَرِّ مَا صُنِعَ لَهُ",
     "Allahumma lakal-hamdu anta kasawtanihi, as'aluka min khayrihi wa khayri ma suni'a lahu, wa a'udhu bika min sharrihi wa sharri ma suni'a lahu",
     "O Allah, all praise is for You. You have dressed me with it. I ask You for the good of it and the good for which it was made, and I seek refuge in You from its evil and the evil for which it was made.",
     "Sunan Abu Dawud", "4020", "HASAN", "daily-life",
     ["grateful"],
     ["clothes", "dressing", "new-clothes", "morning"],
     True, False, True),

    (66, "dua-when-sneezing", "Dua When Sneezing",
     "الْحَمْدُ لِلَّهِ",
     "Alhamdulillah",
     "All praise is for Allah. (The one who hears it responds: Yarhamukallah — May Allah have mercy on you.)",
     "Sahih al-Bukhari", "6224", "SAHIH", "daily-life",
     ["grateful"],
     ["sneezing", "daily", "health"],
     True, False, False),

    (67, "dua-looking-in-mirror", "Dua When Looking in the Mirror",
     "اللَّهُمَّ أَنْتَ حَسَّنْتَ خَلْقِي فَحَسِّنْ خُلُقِي",
     "Allahumma anta hassanta khalqi fahassin khuluqi",
     "O Allah, just as You made my physical form beautiful, make my character beautiful too.",
     "Musnad Ahmad", "3823", "SAHIH", "daily-life",
     ["grateful", "hopeful"],
     ["mirror", "character", "morning", "appearance"],
     True, False, True),

    (68, "dua-entering-market", "Dua When Entering the Market",
     "لَا إِلَهَ إِلَّا اللَّهُ وَحْدَهُ لَا شَرِيكَ لَهُ لَهُ الْمُلْكُ وَلَهُ الْحَمْدُ يُحْيِي وَيُمِيتُ وَهُوَ حَيٌّ لَا يَمُوتُ بِيَدِهِ الْخَيْرُ وَهُوَ عَلَى كُلِّ شَيْءٍ قَدِيرٌ",
     "La ilaha illallahu wahdahu la sharika lahu, lahul-mulku wa lahul-hamdu yuhyi wa yumitu wa huwa hayyun la yamut, biyadihil-khayr wa huwa 'ala kulli shay'in qadir",
     "There is no deity worthy of worship except Allah, alone, without any partner. To Him belongs all dominion and all praise. He gives life and causes death, and He is Ever-Living and will not die. In His hand is all good and He has power over all things.",
     "Sunan al-Tirmidhi", "3428", "HASAN", "daily-life",
     ["grateful"],
     ["market", "shopping", "business", "bazaar"],
     True, False, False),

    (69, "dua-seeing-crescent-moon", "Dua When Seeing the New Crescent Moon",
     "اللَّهُمَّ أَهِلَّهُ عَلَيْنَا بِالْأَمْنِ وَالْإِيمَانِ وَالسَّلَامَةِ وَالْإِسْلَامِ رَبِّي وَرَبُّكَ اللَّهُ",
     "Allahumma ahillahu 'alayna bil-amni wal-iman was-salamati wal-islam, rabbi wa rabbukallah",
     "O Allah, let this moon appear on us with security, faith, safety, and Islam. My Lord and your Lord is Allah.",
     "Sunan al-Tirmidhi", "3451", "HASAN", "daily-life",
     ["grateful", "hopeful"],
     ["moon", "new-month", "ramadan", "dhul-hijjah"],
     True, False, False),

    (70, "dua-leaving-bathroom", "Dua When Leaving the Bathroom",
     "غُفْرَانَكَ",
     "Ghufranaka",
     "I seek Your forgiveness.",
     "Sunan Abu Dawud", "30", "SAHIH", "daily-life",
     ["seeking-forgiveness"],
     ["bathroom", "toilet", "leaving", "daily"],
     True, False, False),

    (71, "dua-for-rizq-and-good-deeds", "Dua for Rizq, Knowledge, and Accepted Deeds",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ عِلْمًا نَافِعًا وَرِزْقًا طَيِّبًا وَعَمَلًا مُتَقَبَّلًا",
     "Allahumma inni as'aluka 'ilman nafi'an wa rizqan tayyiban wa 'amalan mutaqabbala",
     "O Allah, I ask You for beneficial knowledge, good provision, and accepted deeds.",
     "Sunan Ibn Majah", "925", "SAHIH", "daily-life",
     ["hopeful", "grateful"],
     ["rizq", "provision", "knowledge", "work", "morning"],
     True, False, True),

    (72, "dua-for-host-after-meal", "Dua for the Host After Being Hosted",
     "اللَّهُمَّ بَارِكْ لَهُمْ فِيمَا رَزَقْتَهُمْ وَاغْفِرْ لَهُمْ وَارْحَمْهُمْ",
     "Allahumma barik lahum fima razaqtahum waghfir lahum warhamhum",
     "O Allah, bless them in what You have provided for them, forgive them, and have mercy on them.",
     "Sahih Muslim", "2042", "SAHIH", "daily-life",
     ["grateful"],
     ["food", "guest", "hospitality", "blessing", "water", "drink"],
     True, False, False),

    (73, "dua-jazakallahu-khayran", "Dua of Gratitude to Others – Jazakallahu Khayran",
     "جَزَاكَ اللَّهُ خَيْرًا",
     "Jazakallahu khayran",
     "May Allah reward you with good.",
     "Sunan al-Tirmidhi", "2035", "SAHIH", "daily-life",
     ["grateful"],
     ["gratitude", "thanking", "daily", "social"],
     True, False, False),

    (74, "dua-when-it-thunders", "Dua When Hearing Thunder",
     "سُبْحَانَ الَّذِي يُسَبِّحُ الرَّعْدُ بِحَمْدِهِ وَالْمَلَائِكَةُ مِنْ خِيفَتِهِ",
     "Subhanal-ladhi yusabbihur-ra'du bihamdihi wal-mala'ikatu min khifatih",
     "How perfect is He whom the thunder glorifies with His praise, as do the angels out of fear of Him.",
     "Muwatta Imam Malik", "3/154", "HASAN", "daily-life",
     ["afraid", "grateful"],
     ["thunder", "lightning", "storm", "weather"],
     True, False, False),

    (75, "dua-when-wind-blows", "Dua When the Wind Blows",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ خَيْرَهَا وَخَيْرَ مَا فِيهَا وَخَيْرَ مَا أُرْسِلَتْ بِهِ وَأَعُوذُ بِكَ مِنْ شَرِّهَا وَشَرِّ مَا فِيهَا وَشَرِّ مَا أُرْسِلَتْ بِهِ",
     "Allahumma inni as'aluka khayriha wa khayra ma fiha wa khayra ma ursilat bih, wa a'udhu bika min sharriha wa sharri ma fiha wa sharri ma ursilat bih",
     "O Allah, I ask You for the good of it, the good within it, and the good it was sent with. And I seek refuge in You from its evil, the evil within it, and the evil it was sent with.",
     "Sahih Muslim", "899", "SAHIH", "daily-life",
     ["afraid"],
     ["wind", "storm", "weather", "protection"],
     True, False, False),

    (76, "dua-morning-asbahna", "Morning Adhkar – Entering the Morning",
     "أَصْبَحْنَا وَأَصْبَحَ الْمُلْكُ لِلَّهِ وَالْحَمْدُ لِلَّهِ لَا إِلَهَ إِلَّا اللَّهُ وَحْدَهُ لَا شَرِيكَ لَهُ لَهُ الْمُلْكُ وَلَهُ الْحَمْدُ وَهُوَ عَلَى كُلِّ شَيْءٍ قَدِيرٌ",
     "Asbahna wa asbahal-mulku lillah, walhamdu lillah, la ilaha illallahu wahdahu la sharika lahu, lahul-mulku wa lahul-hamdu wa huwa 'ala kulli shay'in qadir",
     "We have entered the morning and the whole kingdom belongs to Allah. All praise is for Allah. There is no deity worthy of worship except Allah, alone, without partner. To Him belongs all dominion and all praise, and He has power over all things.",
     "Sunan Abu Dawud", "5077", "SAHIH", "morning-evening",
     ["grateful"],
     ["morning", "adhkar", "waking-up"],
     True, False, True),

    (77, "dua-subhanallahi-wa-bihamdihi", "Best Dhikr – Subhanallahi wa Bihamdihi",
     "سُبْحَانَ اللَّهِ وَبِحَمْدِهِ",
     "Subhanallahi wa bihamdihi",
     "How perfect is Allah and all praise is for Him. (Recited 100 times in the morning and evening wipes away sins even if they were like foam on the sea.)",
     "Sahih al-Bukhari", "6405", "SAHIH", "morning-evening",
     ["grateful", "seeking-forgiveness"],
     ["morning", "evening", "dhikr", "100-times", "forgiveness", "zikr"],
     True, False, True),

    (78, "dua-protection-from-hellfire", "Dua for Protection from Hellfire",
     "اللَّهُمَّ أَجِرْنِي مِنَ النَّارِ",
     "Allahumma ajirni minan-nar",
     "O Allah, protect me from the Fire.",
     "Sunan Abu Dawud", "5079", "SAHIH", "protection",
     ["afraid", "hopeful", "seeking-forgiveness"],
     ["hellfire", "protection", "morning", "evening", "akhirah"],
     True, False, True),

    (79, "dua-inna-lillahi-upon-calamity", "Dua of Patience Upon Calamity (Inna Lillahi)",
     "إِنَّا لِلَّهِ وَإِنَّا إِلَيْهِ رَاجِعُونَ اللَّهُمَّ أْجُرْنِي فِي مُصِيبَتِي وَأَخْلِفْ لِي خَيْرًا مِنْهَا",
     "Inna lillahi wa inna ilayhi raji'un, Allahumma ajurni fi musibati wa akhlif li khayran minha",
     "Verily we belong to Allah, and truly to Him we shall return. O Allah, reward me for my affliction and replace it for me with something better.",
     "Sahih Muslim", "918", "SAHIH", "protection",
     ["sad", "afraid", "hopeful", "angry"],
     ["death", "calamity", "loss", "grief", "patience"],
     True, False, False),

    (80, "dua-for-guidance-and-righteousness", "Dua for Guidance and Righteousness",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ الْهُدَى وَالتُّقَى وَالْعَفَافَ وَالْغِنَى",
     "Allahumma inni as'alukal-huda wat-tuqa wal-'afafa wal-ghina",
     "O Allah, I ask You for guidance, piety, chastity, and self-sufficiency.",
     "Sahih Muslim", "2721", "SAHIH", "protection",
     ["hopeful", "seeking-forgiveness", "angry", "lonely"],
     ["guidance", "piety", "protection", "daily"],
     True, False, True),

    (81, "dua-severe-hardship-distress", "Dua in Times of Severe Hardship",
     "لَا إِلَهَ إِلَّا اللَّهُ الْحَلِيمُ الْكَرِيمُ سُبْحَانَ اللَّهِ رَبِّ الْعَرْشِ الْعَظِيمِ الْحَمْدُ لِلَّهِ رَبِّ الْعَالَمِينَ",
     "La ilaha illallahul-halimul-karim, subhanallahi rabbil-'arshil-'azim, alhamdu lillahi rabbil-'alamin",
     "There is no deity worthy of worship except Allah, the Forbearing, the Generous. How perfect is Allah, the Lord of the Magnificent Throne. All praise is for Allah, the Lord of all the worlds.",
     "Sahih al-Bukhari", "6346", "SAHIH", "protection",
     ["stressed", "sad", "afraid", "hopeful", "angry"],
     ["hardship", "distress", "difficulty", "anxiety"],
     True, False, True),

    (82, "dua-pain-relief-hand-placement", "Dua for Pain Relief (Hand on Affected Area)",
     "بِسْمِ اللَّهِ بِسْمِ اللَّهِ بِسْمِ اللَّهِ أَعُوذُ بِاللَّهِ وَقُدْرَتِهِ مِنْ شَرِّ مَا أَجِدُ وَأُحَاذِرُ",
     "Bismillah, bismillah, bismillah, a'udhu billahi wa qudratihi min sharri ma ajidu wa uhadhir",
     "In the name of Allah (x3). I seek refuge in Allah and in His power from the evil of what I feel and what I fear.",
     "Sahih Muslim", "2202", "SAHIH", "health",
     ["afraid", "hopeful"],
     ["pain", "illness", "healing", "ruqyah"],
     True, False, False),

    (83, "dua-for-sick-person-seven-times", "Dua for a Sick Person – Recited 7 Times",
     "أَسْأَلُ اللَّهَ الْعَظِيمَ رَبَّ الْعَرْشِ الْعَظِيمِ أَنْ يَشْفِيَكَ",
     "As'alullaha-l-'azima rabba-l-'arshi-l-'azimi an yashfiyak",
     "I ask Allah the Almighty, Lord of the Magnificent Throne, to cure you. (Recited 7 times when visiting a sick Muslim.)",
     "Sunan Abu Dawud", "3106", "SAHIH", "health",
     ["sad", "hopeful"],
     ["visiting-sick", "illness", "healing", "shifa"],
     True, False, False),

    (84, "dua-for-newborn-protection", "Dua of Protection for a Newborn Baby",
     "أُعِيذُكَ بِكَلِمَاتِ اللَّهِ التَّامَّةِ مِنْ كُلِّ شَيْطَانٍ وَهَامَّةٍ وَمِنْ كُلِّ عَيْنٍ لَامَّةٍ",
     "U'idhuka bi kalimatillahit-tammati min kulli shaytanin wa hammatin wa min kulli 'aynin lammah",
     "I seek protection for you in the perfect words of Allah from every devil, every harmful creature, and every evil eye.",
     "Sahih al-Bukhari", "3371", "SAHIH", "family",
     ["afraid", "hopeful"],
     ["newborn", "baby", "children", "protection", "family"],
     True, False, False),

    (85, "dua-before-marital-intimacy", "Dua Before Marital Intimacy",
     "بِسْمِ اللَّهِ اللَّهُمَّ جَنِّبْنَا الشَّيْطَانَ وَجَنِّبِ الشَّيْطَانَ مَا رَزَقْتَنَا",
     "Bismillah, Allahumma jannibna ash-shaytana wa jannibish-shaytana ma razaqtana",
     "In the name of Allah. O Allah, keep Satan away from us and keep Satan away from whatever You bless us with.",
     "Sahih al-Bukhari", "141", "SAHIH", "family",
     ["grateful", "hopeful"],
     ["marriage", "husband-wife", "protection", "family"],
     True, False, False),

    (86, "dua-blessing-at-nikah", "Dua for Blessing at a Wedding (Nikah)",
     "بَارَكَ اللَّهُ لَكَ وَبَارَكَ عَلَيْكَ وَجَمَعَ بَيْنَكُمَا فِي خَيْرٍ",
     "Barakallahu laka wa baraka 'alayk wa jama'a baynakuma fi khayr",
     "May Allah bless you, and shower His blessings upon you, and join you together in goodness.",
     "Sunan Abu Dawud", "2130", "SAHIH", "family",
     ["grateful", "hopeful"],
     ["nikah", "wedding", "marriage", "family", "blessing"],
     True, False, False),

    (87, "dua-visiting-graves", "Dua When Visiting the Graveyard",
     "السَّلَامُ عَلَيْكُمْ أَهْلَ الدِّيَارِ مِنَ الْمُؤْمِنِينَ وَالْمُسْلِمِينَ وَإِنَّا إِنْ شَاءَ اللَّهُ بِكُمْ لَاحِقُونَ نَسْأَلُ اللَّهَ لَنَا وَلَكُمُ الْعَافِيَةَ",
     "Assalamu 'alaykum ahla ad-diyari minal-mu'minina wal-muslimin, wa inna in sha'a Allahu bikum lahiqun, nas'alullaha lana wa lakumul-'afiyah",
     "Peace be upon you, O people of these dwellings, from among the believers and Muslims. We will, if Allah wills, be joining you. We ask Allah for well-being for us and for you.",
     "Sahih Muslim", "974", "SAHIH", "family",
     ["sad", "hopeful"],
     ["graves", "cemetery", "deceased", "death"],
     True, False, False),

    (88, "dua-janazah-prayer", "Dua for the Deceased in Janazah Prayer",
     "اللَّهُمَّ اغْفِرْ لِحَيِّنَا وَمَيِّتِنَا وَشَاهِدِنَا وَغَائِبِنَا وَصَغِيرِنَا وَكَبِيرِنَا وَذَكَرِنَا وَأُنْثَانَا",
     "Allahummaghfir lihayyina wa mayyitina wa shahidina wa gha'ibina wa saghirina wa kabirina wa dhakarana wa unthana",
     "O Allah, forgive our living and our dead, those who are present and those who are absent, our young and our old, our males and our females.",
     "Sunan Abu Dawud", "3201", "SAHIH", "family",
     ["sad", "hopeful"],
     ["janazah", "funeral", "deceased", "death"],
     True, False, False),

    (89, "dua-to-make-journey-easy", "Dua to Make the Journey Easy",
     "اللَّهُمَّ هَوِّنْ عَلَيْنَا سَفَرَنَا هَذَا وَاطْوِ عَنَّا بُعْدَهُ",
     "Allahumma hawwin 'alayna safarana hadha wattwi 'anna bu'dah",
     "O Allah, make this journey of ours easy for us and fold up its distance for us.",
     "Sahih Muslim", "1342", "SAHIH", "travel",
     ["hopeful"],
     ["travel", "journey", "car", "plane", "easy", "umrah", "hajj", "pilgrimage"],
     True, False, False),

    (90, "dua-entering-new-town", "Dua When Entering a New City or Town",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ خَيْرَهَا وَخَيْرَ أَهْلِهَا وَخَيْرَ مَا فِيهَا وَأَعُوذُ بِكَ مِنْ شَرِّهَا وَشَرِّ أَهْلِهَا وَشَرِّ مَا فِيهَا",
     "Allahumma inni as'aluka khayriha wa khayra ahliha wa khayra ma fiha, wa a'udhu bika min sharriha wa sharri ahliha wa sharri ma fiha",
     "O Allah, I ask You for the good of it, the good of its people, and the good of what is in it. And I seek refuge in You from its evil, the evil of its people, and the evil of what is in it.",
     "Sahih al-Bukhari", "2893", "SAHIH", "travel",
     ["afraid", "hopeful"],
     ["travel", "new-place", "city", "town", "protection", "umrah", "hajj", "new-city"],
     True, False, False),

    (91, "dua-seeking-forgiveness-100-times", "Dua of Istighfar – Seeking Forgiveness",
     "اللَّهُمَّ اغْفِرْ لِي وَتُبْ عَلَيَّ إِنَّكَ أَنْتَ التَّوَّابُ الْغَفُورُ",
     "Allahumma aghfir li wa tub 'alayya innaka antat-tawwabul-ghafur",
     "O Allah, forgive me and accept my repentance. Indeed You are the Oft-Returning, the Most Forgiving. (The Prophet recited this 100 times per day.)",
     "Sunan Ibn Majah", "3814", "SAHIH", "forgiveness",
     ["seeking-forgiveness", "hopeful"],
     ["istighfar", "forgiveness", "repentance", "100-times"],
     True, False, True),

    (92, "dua-when-fasting-provoked", "What to Say When Fasting and Provoked",
     "إِنِّي صَائِمٌ إِنِّي صَائِمٌ",
     "Inni sa'im, inni sa'im",
     "I am fasting, I am fasting. (Said aloud or to oneself when someone provokes you while fasting.)",
     "Sahih al-Bukhari", "1904", "SAHIH", "worship",
     ["angry"],
     ["fasting", "ramadan", "anger", "provoked"],
     True, False, False),

    (93, "dua-seeking-rain-istisqa", "Dua for Rain (Istisqa – Seeking Rainfall)",
     "اللَّهُمَّ أَسْقِنَا غَيْثًا مُغِيثًا مَرِيئًا مَرِيعًا نَافِعًا غَيْرَ ضَارٍّ عَاجِلًا غَيْرَ آجِلٍ",
     "Allahumma asqina ghayathan mughithan mari'an mari'an nafi'an ghayra darrin 'ajilan ghayra ajil",
     "O Allah, send upon us rain that is beneficial, wholesome, abundant, and nourishing — not harmful — and soon, not delayed.",
     "Sunan Abu Dawud", "1169", "SAHIH", "daily-life",
     ["hopeful"],
     ["rain", "drought", "istisqa", "weather"],
     True, False, False),

    (94, "dua-salawat-ibrahim", "Salawat Ibrahimiyyah – Full Blessings on the Prophet",
     "اللَّهُمَّ صَلِّ عَلَى مُحَمَّدٍ وَعَلَى آلِ مُحَمَّدٍ كَمَا صَلَّيْتَ عَلَى إِبْرَاهِيمَ وَعَلَى آلِ إِبْرَاهِيمَ إِنَّكَ حَمِيدٌ مَجِيدٌ وَبَارِكْ عَلَى مُحَمَّدٍ وَعَلَى آلِ مُحَمَّدٍ كَمَا بَارَكْتَ عَلَى إِبْرَاهِيمَ وَعَلَى آلِ إِبْرَاهِيمَ إِنَّكَ حَمِيدٌ مَجِيدٌ",
     "Allahumma salli 'ala Muhammadin wa 'ala ali Muhammadin kama sallayta 'ala Ibrahima wa 'ala ali Ibrahim, innaka Hamidum Majid. Wa barik 'ala Muhammadin wa 'ala ali Muhammadin kama barakta 'ala Ibrahima wa 'ala ali Ibrahim, innaka Hamidum Majid",
     "O Allah, send blessings upon Muhammad and upon the family of Muhammad, as You sent blessings upon Ibrahim and upon the family of Ibrahim. Indeed, You are Praiseworthy and Glorious. And send peace upon Muhammad and upon the family of Muhammad, as You sent peace upon Ibrahim and upon the family of Ibrahim. Indeed, You are Praiseworthy and Glorious.",
     "Sahih al-Bukhari", "3370", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["salawat", "prophet", "prayer", "friday", "tashahhud", "jummah"],
     True, False, True),

    (95, "dua-seeing-someone-in-difficulty", "Dua When Seeing Someone in Affliction",
     "الْحَمْدُ لِلَّهِ الَّذِي عَافَانِي مِمَّا ابْتَلَاهُ بِهِ وَفَضَّلَنِي عَلَى كَثِيرٍ مِمَّنْ خَلَقَ تَفْضِيلًا",
     "Alhamdu lillahil-ladhi 'afani mimma-btalahu bihi wa fadddalani 'ala kathirin mimman khalaqa tafdhila",
     "All praise is for Allah who saved me from that which He has afflicted you with, and Who has truly favoured me over many of those He has created.",
     "Sunan al-Tirmidhi", "3431", "HASAN", "daily-life",
     ["grateful"],
     ["affliction", "gratitude", "illness", "hardship"],
     True, False, False),

    (96, "dua-for-strong-faith-completion", "Dua for Firm Faith and Good End",
     "اللَّهُمَّ يَا مُقَلِّبَ الْقُلُوبِ وَالْأَبْصَارِ ثَبِّتْ قَلْبِي عَلَى دِينِكَ",
     "Allahumma ya muqallibal-qulubi wal-absar, thabbit qalbi 'ala dinik",
     "O Allah, O Turner of hearts and sight, make my heart firm upon Your religion.",
     "Sunan al-Tirmidhi", "2066", "SAHIH", "protection",
     ["afraid", "hopeful", "lonely"],
     ["faith", "iman", "heart", "steadfastness", "death"],
     True, False, True),

    (97, "dua-for-sleep-protection-night", "Dua Before Sleep for Nightly Protection",
     "اللَّهُمَّ قِنِي عَذَابَكَ يَوْمَ تَبْعَثُ عِبَادَكَ",
     "Allahumma qini 'adhabaka yawma tab'athu 'ibadak",
     "O Allah, protect me from Your punishment on the Day You resurrect Your servants.",
     "Sunan Abu Dawud", "5045", "SAHIH", "daily-life",
     ["afraid", "hopeful", "lonely"],
     ["before-sleep", "night", "bedtime", "protection", "akhirah", "sleep", "sleeping"],
     True, False, True),

    (98, "dua-for-complete-wellbeing-afiyah", "Dua for Complete Wellbeing ('Afiyah)",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ الْعَفْوَ وَالْعَافِيَةَ فِي الدُّنْيَا وَالْآخِرَةِ",
     "Allahumma inni as'alukal-'afwa wal-'afiyata fid-dunya wal-akhirah",
     "O Allah, I ask You for pardon and well-being in this world and in the Hereafter.",
     "Sunan Ibn Majah", "3871", "SAHIH", "health",
     ["hopeful", "grateful"],
     ["wellbeing", "health", "pardon", "daily", "morning"],
     True, False, True),

    (99, "dua-for-those-who-show-kindness", "Dua Asking Allah to Reward Those Who Show Kindness",
     "اللَّهُمَّ أَعِنِّي عَلَى شُكْرِكَ وَذِكْرِكَ وَحُسْنِ عِبَادَتِكَ",
     "Allahumma a'inni 'ala shukrika wa dhikrika wa husni 'ibadatik",
     "O Allah, help me to be grateful to You, to remember You, and to worship You in the best manner.",
     "Sunan Abu Dawud", "1522", "SAHIH", "worship",
     ["grateful", "hopeful"],
     ["gratitude", "dhikr", "worship", "daily", "zikr"],
     True, False, True),

    (100, "dua-for-seeking-allahs-pleasure", "Dua for Allah's Pleasure and Refuge from His Anger",
     "اللَّهُمَّ إِنِّي أَسْأَلُكَ رِضَاكَ وَالْجَنَّةَ وَأَعُوذُ بِكَ مِنْ سَخَطِكَ وَالنَّارِ",
     "Allahumma inni as'aluka ridaka wal-jannah, wa a'udhu bika min sakhatika wan-nar",
     "O Allah, I ask You for Your pleasure and Paradise, and I seek refuge in You from Your anger and the Fire.",
     "Sunan Ibn Majah", "3846", "SAHIH", "worship",
     ["hopeful", "afraid", "seeking-forgiveness"],
     ["jannah", "paradise", "akhirah", "pleasure", "daily"],
     True, False, True),
]

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    ("#",                    5),
    ("Title (English)",     28),
    ("Arabic Text",         35),
    ("Transliteration",     35),
    ("Translation (English Meaning)", 45),
    ("Source Book",         22),
    ("Hadith / Reference Number", 12),
    ("Authenticity Grade",  14),
    ("Category",            18),
    ("Emotion Tags",        25),
    ("Situation Tags",      30),
    ("Scholar Verified",    14),
    ("Live on Website URL", 50),
    ("Scholar Notes",       35),
    ("Correction Required?",18),
]

BASE_URL = "https://dua-companion-rho.vercel.app/duas/"

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1d4e38")   # dark emerald
ALT_ROW_FILL  = PatternFill("solid", fgColor="f0faf5")   # very light green
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Arial", size=10)
ARABIC_FONT   = Font(name="Arial", size=13)   # larger for Arabic

WRAP          = Alignment(wrap_text=True, vertical="top")
WRAP_RIGHT    = Alignment(wrap_text=True, vertical="top",
                           horizontal="right", reading_order=2)  # RTL
CENTER        = Alignment(horizontal="center", vertical="center")

THIN_BORDER   = Border(
    left=Side(style="thin",   color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
    top=Side(style="thin",    color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

GRADE_COLORS = {
    "SAHIH":  ("D4EDDA", "155724"),   # green-ish
    "HASAN":  ("FFF3CD", "856404"),   # yellow-ish
    "QURAN":  ("CCE5FF", "004085"),   # blue-ish
    "DA'IF":  ("F8D7DA", "721C24"),   # red-ish
}


def grade_fill(grade: str, row_is_alt: bool):
    colors = GRADE_COLORS.get(grade)
    if colors:
        return PatternFill("solid", fgColor=colors[0])
    return ALT_ROW_FILL if row_is_alt else WHITE_FILL


def grade_font(grade: str):
    colors = GRADE_COLORS.get(grade)
    if colors:
        return Font(name="Arial", size=10, color=colors[1], bold=True)
    return BODY_FONT


# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Duas for Review"

# ── Write header ──────────────────────────────────────────────────────────────
for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"

# ── Write data rows ───────────────────────────────────────────────────────────
ARABIC_COL   = 3   # 1-based column index of Arabic Text
GRADE_COL    = 8   # authenticity grade column

for row_num, dua in enumerate(DUAS, start=2):
    (dua_id, slug, title, arabic, translit, translation,
     source_book, hadith_num, grade, category,
     emotions, situations, scholar_verified, featured, daily_eligible) = dua

    is_alt = (row_num % 2 == 0)
    default_fill = ALT_ROW_FILL if is_alt else WHITE_FILL

    url = BASE_URL + slug
    row_data = [
        dua_id,
        title,
        arabic,
        translit,
        translation,
        source_book,
        hadith_num,
        grade,
        category,
        ", ".join(emotions),
        ", ".join(situations),
        "Yes" if scholar_verified else "No",
        url,
        "",   # Scholar Notes – blank
        "",   # Correction Required? – blank
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = THIN_BORDER

        # Arabic text column
        if col_idx == ARABIC_COL:
            cell.font      = ARABIC_FONT
            cell.alignment = WRAP_RIGHT
            cell.fill      = default_fill
        # Grade column – colour-coded
        elif col_idx == GRADE_COL:
            cell.font      = grade_font(str(value))
            cell.fill      = grade_fill(str(value), is_alt)
            cell.alignment = Alignment(horizontal="center", vertical="top",
                                       wrap_text=True)
        else:
            cell.font      = BODY_FONT
            cell.alignment = WRAP
            cell.fill      = default_fill

    ws.row_dimensions[row_num].height = 60

# ── Legend sheet ──────────────────────────────────────────────────────────────
legend = wb.create_sheet(title="Legend")

legend_title_font = Font(name="Arial", bold=True, size=13, color="1d4e38")
legend_h2_font    = Font(name="Arial", bold=True, size=11, color="FFFFFF")
legend_body_font  = Font(name="Arial", size=10)
legend_bold_font  = Font(name="Arial", bold=True, size=10)

legend.column_dimensions["A"].width = 28
legend.column_dimensions["B"].width = 70

def legend_section_header(row, text):
    cell = legend.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    legend.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=2)
    legend.row_dimensions[row].height = 24

def legend_row(row, label, description, bold_label=True):
    a = legend.cell(row=row, column=1, value=label)
    a.font = legend_bold_font if bold_label else legend_body_font
    a.alignment = Alignment(wrap_text=True, vertical="top")

    b = legend.cell(row=row, column=2, value=description)
    b.font = legend_body_font
    b.alignment = Alignment(wrap_text=True, vertical="top")
    legend.row_dimensions[row].height = 36

# Title
title_cell = legend.cell(row=1, column=1,
    value="Dua Companion – Scholar Review File Legend")
title_cell.font = legend_title_font
legend.merge_cells("A1:B1")
legend.row_dimensions[1].height = 30

# ── Section 1: Authenticity Grades ───────────────────────────────────────────
legend_section_header(3, "AUTHENTICITY GRADES")

grades = [
    ("SAHIH",  "Authentic – Narration meets all conditions of sound transmission. "
               "Highest grade of hadith authenticity."),
    ("HASAN",  "Good – Narration is slightly below Sahih in chain strength but still "
               "acceptable for practice."),
    ("QURAN",  "Quranic Verse – Direct ayah from the Quran. Absolute authority in "
               "Islamic practice."),
    ("DA'IF",  "Weak – Contains a deficiency in the chain or text. Scholars differ on "
               "using weak narrations for supplication (du'a). Flagged for review."),
]
for i, (g, desc) in enumerate(grades, start=4):
    legend_row(i, g, desc)
    # colour code grade label
    fill_hex, font_hex = GRADE_COLORS.get(g, ("FFFFFF", "000000"))
    legend.cell(row=i, column=1).fill = PatternFill("solid", fgColor=fill_hex)
    legend.cell(row=i, column=1).font = Font(
        name="Arial", bold=True, size=10, color=font_hex)

# ── Section 2: Column Descriptions ───────────────────────────────────────────
legend_section_header(9, "COLUMN DESCRIPTIONS")

col_descriptions = [
    ("#",                    "Sequential ID number for each dua."),
    ("Title (English)",      "Human-readable name describing the context or purpose "
                             "of the dua."),
    ("Arabic Text",          "The original Arabic text of the dua, right-aligned "
                             "(RTL reading order)."),
    ("Transliteration",      "Romanised phonetic representation of the Arabic to aid "
                             "non-Arabic readers in pronunciation."),
    ("Translation",          "English meaning of the dua for comprehension and context."),
    ("Source Book",          "The primary hadith collection or scripture (Quran) from "
                             "which the dua is sourced."),
    ("Hadith / Ref Number",  "Specific hadith number or Quranic verse reference "
                             "(e.g. 2:201 = Surah 2, Ayah 201)."),
    ("Authenticity Grade",   "Classification of the narration's strength: SAHIH, "
                             "HASAN, QURAN, or DA'IF (see Grades section above)."),
    ("Category",             "Primary topical category assigned to the dua on the "
                             "website (e.g. daily-life, worship, protection, family)."),
    ("Emotion Tags",         "Emotional states this dua is recommended for "
                             "(comma-separated list)."),
    ("Situation Tags",       "Situational keywords used for search and filtering on "
                             "the platform (comma-separated list)."),
    ("Scholar Verified",     "Whether the dua has been reviewed by a qualified Islamic "
                             "scholar prior to publication."),
    ("Live on Website URL",  "Direct URL to the dua's page on Dua Companion. "
                             "Open to view how it currently appears to users."),
    ("Scholar Notes",        "FREE-FORM FIELD for scholars to add comments, "
                             "corrections, references, or additional context."),
    ("Correction Required?", "Enter YES if a correction is needed, NO if everything "
                             "is accurate."),
]
for i, (col, desc) in enumerate(col_descriptions, start=10):
    legend_row(i, col, desc)

# ── Section 3: Filling Instructions ──────────────────────────────────────────
legend_section_header(26, "INSTRUCTIONS FOR SCHOLARS")

instructions = [
    ("Step 1 – Review",
     "Read each dua's Arabic text, transliteration, and translation carefully. "
     "Verify correctness against the cited source."),
    ("Step 2 – Scholar Notes",
     "In the 'Scholar Notes' column, write any observations. Examples: "
     "'Translation slightly inaccurate – suggest rewording…', "
     "'Hadith number should be 6325 not 6324', "
     "'Consider adding context about when this is recited'."),
    ("Step 3 – Correction Required?",
     "Mark 'Yes' if any part of the dua entry needs correction (Arabic, "
     "transliteration, translation, source, grade, or tags). "
     "Mark 'No' if the entry is accurate and ready to remain live."),
    ("Step 4 – Arabic Text",
     "If the Arabic text itself needs correction, please write the corrected "
     "Arabic in the Scholar Notes column clearly labelled: "
     "'Corrected Arabic: [text here]'."),
    ("Step 5 – Grade Disputes",
     "If you believe the Authenticity Grade is incorrect, note the correct grade "
     "and supporting reference in Scholar Notes."),
    ("Contact",
     "Return the completed file to the Dua Companion team. "
     "All corrections will be reviewed and applied to the live platform."),
]
for i, (step, text) in enumerate(instructions, start=27):
    legend_row(i, step, text)

# ── Save file ─────────────────────────────────────────────────────────────────
OUTPUT_PATH = "/Users/mohdtalhamasood/Downloads/Duas_Scholar_Review.xlsx"
wb.save(OUTPUT_PATH)

print(f"SUCCESS: Excel file created at {OUTPUT_PATH}")
print(f"Total duas included: {len(DUAS)}")
print(f"Sheets: '{ws.title}', 'Legend'")
